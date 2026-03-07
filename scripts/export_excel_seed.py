from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/Users/nicole/Desktop/需求价值测算_团队版.xlsx")
OUTPUT = ROOT / "data" / "workbook-seed.json"


def score_roi(raw_roi: float) -> int:
    if raw_roi >= 3000:
        return 5
    if raw_roi >= 2000:
        return 4
    if raw_roi >= 1000:
        return 3
    if raw_roi >= 500:
        return 2
    return 1


def category(score: float) -> str:
    if score >= 4:
      return "A类（优先内部）"
    if score >= 3.5:
      return "B类（保留优化）"
    if score >= 2.7:
      return "C类（可外包）"
    return "D类（淘汰）"


def created_at_for_demo_row(row_index: int) -> str:
    if 24 <= row_index <= 114:
        month = "2025-11"
        day = min(row_index - 23, 28)
    elif 145 <= row_index <= 294:
        month = "2025-12"
        day = ((row_index - 145) % 28) + 1
    elif 295 <= row_index <= 363:
        month = "2026-01"
        day = min(row_index - 294, 28)
    else:
        month = "2025-12"
        day = ((row_index - 24) % 28) + 1

    return f"{month}-{day:02d}T00:00:00.000Z"


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=True)

    price_sheet = workbook["报价表"]
    pricing_options = []
    for row in price_sheet.iter_rows(min_row=2, values_only=True):
        label = row[0]
        if not label:
            continue

        prices = {}
        for key, index in [("插画+动效", 1), ("仅动效", 2), ("仅插画", 3), ("三维设计", 4)]:
            value = row[index]
            if isinstance(value, (int, float)):
                prices[key] = value

        pricing_options.append({"label": label, "prices": prices})

    designer_sheet = workbook["Designer_Efficiency"]
    designers = []
    for row in designer_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        designers.append({"name": row[0], "efficiency": float(row[1] or 1)})

    revision_coefficients = [
        {"round": 0, "coefficient": 1.0},
        {"round": 1, "coefficient": 0.9},
        {"round": 2, "coefficient": 0.8},
        {"round": 3, "coefficient": 0.65},
        {"round": 4, "coefficient": 0.5},
        {"round": 5, "coefficient": 0.35},
    ]

    demo_sheet = workbook["demo"]
    sample_records = []
    breakdown = Counter()

    efficiency_by_type = {
        "插画+动效": 1.2,
        "仅动效": 1.1,
        "仅插画": 1.0,
        "三维设计": 1.3,
    }

    for row_index in range(24, 364):
        business_line = demo_sheet.cell(row_index, 1).value
        demand_label = demo_sheet.cell(row_index, 2).value
        pricing_type = demo_sheet.cell(row_index, 7).value
        base_value = demo_sheet.cell(row_index, 8).value
        designer = demo_sheet.cell(row_index, 9).value
        efficiency = demo_sheet.cell(row_index, 10).value
        avg_days = demo_sheet.cell(row_index, 11).value
        slicing_days = demo_sheet.cell(row_index, 12).value
        total_cost = demo_sheet.cell(row_index, 13).value
        revision_rounds = demo_sheet.cell(row_index, 14).value
        revision_coefficient = demo_sheet.cell(row_index, 15).value
        raw_roi = demo_sheet.cell(row_index, 16).value
        strategic_value = demo_sheet.cell(row_index, 17).value
        composite_score = demo_sheet.cell(row_index, 18).value
        suggested = demo_sheet.cell(row_index, 19).value
        confidence_score = demo_sheet.cell(row_index, 20).value

        if not (business_line and demand_label and pricing_type and base_value is not None):
            continue

        efficiency_value = float(efficiency or efficiency_by_type.get(str(pricing_type), 1))
        revision_rounds_value = int(float(revision_rounds or 0))
        revision_coefficient_value = float(
            revision_coefficient
            if revision_coefficient is not None
            else next(
                (item["coefficient"] for item in revision_coefficients if item["round"] == revision_rounds_value),
                0.35 if revision_rounds_value >= 5 else 1.0,
            )
        )
        total_cost_value = float(total_cost if total_cost is not None else (float(avg_days or 0) + float(slicing_days or 0) + revision_rounds_value))
        raw_roi_value = float(raw_roi if raw_roi is not None else 0)
        roi_score = score_roi(raw_roi_value)
        composite_score_value = float(composite_score if composite_score is not None else 0)
        suggested_value = str(suggested or category(composite_score_value))
        breakdown[suggested_value] += 1

        selected_price = next(
            (
                option["prices"].get(pricing_type)
                for option in pricing_options
                if option["label"] == demand_label
            ),
            float(base_value or 0),
        )

        sample_records.append(
            {
                "id": f"seed-{row_index}",
                "createdAt": created_at_for_demo_row(row_index),
                "businessLine": str(business_line),
                "demandLabel": str(demand_label),
                "pricingType": str(pricing_type),
                "baseValue": float(base_value or 0),
                "designer": str(designer or ""),
                "efficiency": efficiency_value,
                "avgProductionDays": float(avg_days or 0),
                "slicingDays": float(slicing_days or 0),
                "revisionRounds": revision_rounds_value,
                "strategicValue": int(strategic_value or 0),
                "confidenceScore": int(confidence_score or 3),
                "selectedPrice": float(selected_price or 0),
                "totalCost": round(total_cost_value, 2),
                "rawRoi": round(raw_roi_value, 2),
                "roiScore": roi_score,
                "compositeScore": round(composite_score_value, 3),
                "suggestedCategory": suggested_value,
                "revisionCoefficient": revision_coefficient_value,
            }
        )

    payload = {
        "pricingOptions": pricing_options,
        "designers": designers,
        "revisionCoefficients": revision_coefficients,
        "sampleRecords": sample_records,
        "summaries": {
            "totalPricingOptions": len(pricing_options),
            "totalDesigners": len(designers),
            "totalSamples": len(sample_records),
            "categoryBreakdown": dict(breakdown),
        },
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUTPUT}")


if __name__ == "__main__":
    main()
