from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/Users/nicole/Desktop/wowchat产能定编模型_design .xlsx")
OUTPUT = ROOT / "data" / "workforce-seed.json"


def normalize_month(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    if text.isdigit():
        return f"{text}月"
    return text


def number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return 0.0


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=True)

    config_sheet = workbook["配置"]
    monthly_sheet = workbook["汇总_月总"]
    business_sheet = workbook["业务输入"]
    business_list_sheet = workbook["业务清单"]

    safe_utilization = number(config_sheet["B3"].value)
    illustration_headcount = number(config_sheet["B4"].value)
    motion_headcount = number(config_sheet["B5"].value)
    combo_split_illustration = number(config_sheet["B6"].value)
    combo_split_motion = number(config_sheet["B7"].value)
    average_monthly_total = number(config_sheet["B10"].value)
    total_headcount = number(config_sheet["B11"].value)
    actual_per_capita_monthly = number(config_sheet["B12"].value)
    baseline_per_capita_monthly = number(config_sheet["B13"].value)

    config = {
        "safeUtilization": safe_utilization,
        "illustrationHeadcount": illustration_headcount,
        "motionHeadcount": motion_headcount,
        "comboSplitIllustration": combo_split_illustration,
        "comboSplitMotion": combo_split_motion,
        "averageMonthlyTotal": average_monthly_total,
        "totalHeadcount": total_headcount,
        "actualPerCapitaMonthly": actual_per_capita_monthly,
        "baselinePerCapitaMonthly": baseline_per_capita_monthly,
    }

    monthly_totals = []
    month_headers = [normalize_month(monthly_sheet.cell(3, column).value) for column in range(2, 5)]
    total_row = 4
    illustration_row = 5
    motion_row = 6
    unmapped_row = 8

    for index, month in enumerate(month_headers, start=2):
        if not month:
            continue
        monthly_totals.append(
            {
                "month": month,
                "totalAmount": number(monthly_sheet.cell(total_row, index).value),
                "illustrationAmount": number(monthly_sheet.cell(illustration_row, index).value),
                "motionAmount": number(monthly_sheet.cell(motion_row, index).value),
                "unmappedRows": int(number(monthly_sheet.cell(unmapped_row, index).value)),
            }
        )

    business_lines = []
    for row in range(4, business_list_sheet.max_row + 1):
        name = business_list_sheet.cell(row, 1).value
        if name:
            business_lines.append(str(name).strip())

    business_inputs = []
    for row in range(5, business_sheet.max_row + 1):
        business_line = business_sheet.cell(row, 1).value
        if not business_line:
            continue

        month = normalize_month(business_sheet.cell(row, 2).value)
        illustration_amount = number(business_sheet.cell(row, 3).value)
        motion_amount = number(business_sheet.cell(row, 4).value)

        business_inputs.append(
            {
                "businessLine": str(business_line).strip(),
                "month": month,
                "illustrationAmount": illustration_amount,
                "motionAmount": motion_amount,
                "totalAmount": number(business_sheet.cell(row, 5).value) or illustration_amount + motion_amount,
                "concentrationFactor": number(business_sheet.cell(row, 6).value) or 1.0,
                "qualityFactor": number(business_sheet.cell(row, 7).value) or 1.0,
            }
        )

    payload = {
        "config": config,
        "monthlyTotals": monthly_totals,
        "businessInputs": business_inputs,
        "defaultBusinessLine": business_lines[0] if business_lines else "",
        "businessLines": business_lines,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUTPUT}")


if __name__ == "__main__":
    main()
